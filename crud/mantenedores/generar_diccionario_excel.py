import ast
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def extraer_diccionario_desde_modelos(path_archivo):
    """
    Extrae la estructura de modelos desde un archivo diccionario.py generado por inspectdb.
    Devuelve una lista de modelos con sus campos y relaciones.
    """
    # Intentar leer con UTF-8, si falla usar UTF-16
    try:
        with open(path_archivo, "r", encoding="utf-8-sig") as f:
            contenido = f.read()
    except UnicodeDecodeError:
        with open(path_archivo, "r", encoding="utf-16") as f:
            contenido = f.read()

    tree = ast.parse(contenido)
    modelos = []

    for nodo in tree.body:
        if isinstance(nodo, ast.ClassDef):
            nombre_modelo = nodo.name
            tabla = None
            campos = []

            # Buscar db_table dentro de Meta
            for subnodo in nodo.body:
                if isinstance(subnodo, ast.ClassDef) and subnodo.name == "Meta":
                    for meta_linea in subnodo.body:
                        if (
                            isinstance(meta_linea, ast.Assign)
                            and isinstance(meta_linea.targets[0], ast.Name)
                            and meta_linea.targets[0].id == "db_table"
                        ):
                            if isinstance(meta_linea.value, ast.Constant):
                                tabla = meta_linea.value.value

                # Extraer campos
                if isinstance(subnodo, ast.Assign):
                    if isinstance(subnodo.value, ast.Call) and isinstance(subnodo.value.func, ast.Attribute):
                        tipo = subnodo.value.func.attr
                        rel = None
                        if tipo == "ForeignKey" and len(subnodo.value.args) > 0:
                            arg = subnodo.value.args[0]
                            if isinstance(arg, ast.Name):
                                rel = arg.id
                            elif isinstance(arg, ast.Constant):
                                rel = arg.value

                        campos.append({
                            "Campo": subnodo.targets[0].id,
                            "Tipo": tipo,
                            "Relación": rel
                        })

            modelos.append({
                "Modelo": nombre_modelo,
                "Tabla BD": tabla or "",
                "Campos": campos
            })

    return modelos


def exportar_diccionario(modelos):
    """
    Genera un archivo Excel con el diccionario de datos de los modelos extraídos
    con formato profesional y diferenciación por modelo.
    """
    filas = []
    for modelo in modelos:
        for campo in modelo["Campos"]:
            filas.append({
                "Modelo": modelo["Modelo"],
                "Tabla BD": modelo["Tabla BD"],
                "Campo": campo["Campo"],
                "Tipo": campo["Tipo"],
                "Relación": campo["Relación"] or ""
            })

    df = pd.DataFrame(filas)

    # Crear carpeta de salida si no existe
    out_dir = Path("documentacion")
    out_dir.mkdir(exist_ok=True)
    path = out_dir / "diccionario_datos.xlsx"

    df.to_excel(path, index=False)

    # Aplicar formato con openpyxl
    wb = load_workbook(path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Colores alternos por modelo
    modelo_actual = None
    color_fila_1 = "D9E1F2"  # azul claro
    color_fila_2 = "FFFFFF"  # blanco
    color = color_fila_1

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        # Encabezado
        if row[0].row == 1:
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = thin_border
        else:
            # Cambiar color si cambia el modelo
            if row[0].value != modelo_actual:
                modelo_actual = row[0].value
                color = color_fila_1 if color == color_fila_2 else color_fila_2

            for cell in row:
                cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = alignment
                cell.border = thin_border

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    wb.save(path)
    print(f"✅ Diccionario generado con éxito y con diferenciación por modelo: {path.resolve()}")


if __name__ == "__main__":
    archivo_modelos = "diccionario.py"
    if not Path(archivo_modelos).exists():
        print(f"❌ Archivo '{archivo_modelos}' no encontrado. Colócalo en la misma carpeta que este script.")
    else:
        modelos = extraer_diccionario_desde_modelos(archivo_modelos)
        exportar_diccionario(modelos)
